
import pandas as pd
import requests
from bs4 import BeautifulSoup
import pymysql
import os
import logging
import datetime as dt


#this method with read historical VIX data from downloaded excel files and store it to db
def store_exceldata_to_db(files_path, db_hostname, db_username, db_pwd,logging_directory):
    try:
        files = []
        with os.scandir(files_path) as listOfEntries:  
            for entry in listOfEntries:
                if entry.is_file():
                    files.append(entry.name)
        
        con =connect_to_db(db_hostname, db_username, db_pwd)
        
        for i in files:
            file = files_path + "\\" + i
            df = pd.read_excel(file)
            #df.columns = df.columns.str.replace(" ","_")
            df.columns = ['Trade_Date','Contract_Month','Expiration_Date','VIX']
            print(df)
            df['Trade_Date'] = df['Trade_Date'].apply(lambda x : x[0:10])
            df['Trade_Date'] = df['Trade_Date'].apply(lambda x : dt.datetime.strptime(x, ('%Y-%m-%d')).strftime('%Y/%m/%d'))
            df['Expiration_Date'] = df['Expiration_Date'].apply(lambda x : dt.datetime.strptime(x, ('%m/%d/%Y')).strftime('%b-%y'))
            
        
            con =connect_to_db(db_hostname, db_username, db_pwd)
            df.to_sql(name='vix_data',con=con,if_exists='append',index=False,flavor='mysql')
            con.commit()
            con.close()
    
    except Exception as e:
        log_error(logging_directory,"store_exceldata_to_db",e)
    

#get daily vix data from CBOE and store to db    
def get_daily_vix(db_hostname, db_username, db_pwd,logging_directory,
                  web_path ="http://www.cboe.com/trading-tools/strategy-planning-tools/term-structure-data"):
    
    try:
        
        res = requests.get(web_path)
        soup = BeautifulSoup(res.content,'lxml')
        table = soup.find_all('table')[0] 
        df = pd.read_html(str(table))
        df=df[0]
        df.columns =  df.ix[0]
        df = df.ix[1:,:]
        df.columns = df.columns.str.replace(" ","_")
        
        
        df['Trade_Date'] = df['Trade_Date'].apply(lambda x : x[0:9])
        df['Trade_Date'] = df['Trade_Date'].apply(lambda x : dt.datetime.strptime(x, ('%m/%d/%Y')).strftime('%Y/%m/%d'))
        df['Expiration_Date'] = df['Expiration_Date'].apply(lambda x : dt.datetime.strptime(x, ('%d-%b-%y')).strftime('%b-%y'))
        
        con =connect_to_db(db_hostname, db_username, db_pwd)
        df.to_sql(name='vix_data',con=con,if_exists='append',index=False,flavor='mysql')
        con.commit()
        con.close()
        
    except Exception as e:
        log_error(logging_directory,"get_daily_vix",e)
        

#set up database connection and return connection string
def connect_to_db(db_hostname, db_username, db_pwd):
    
    con = pymysql.connect(db_hostname,db_username,db_pwd)
    con.cursor().execute("CREATE DATABASE IF NOT EXISTS vix")
    con.cursor().execute("USE vix")
    
    return con

#logs error in log file on exception
def log_error(logging_directory,calling_method, exception):
    
    date = dt.datetime.now().strftime("%Y_%m_%d")
    logfile=logging_directory+'\\exception'+date+'.log'
    logging.basicConfig(filename=logfile,level=logging.DEBUG)
    logging.warning("caught exception in method: "+calling_method )
    logging.warning("Error: "+str(exception))
    
#get all the data from DB to dataframe
def get_data_from_db(db_hostname, db_username, db_pwd,db_name,table_name):
    
    con = connect_to_db(db_hostname, db_username, db_pwd)
    cursor = con.cursor()
    cursor.execute("SELECT * FROM " +db_name+"."+table_name + " order by trade_date asc,Contract_Month asc")
    result = pd.DataFrame(list(cursor.fetchall()))
    
    return result


#will extract the data from database and save to excel in required format as per Ben's mail
def save_vix_data_to_excel(result = get_data_from_db("localhost","root","test123",'vix','vix_data')):
    x= result.ix[:,0].unique()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    ws['A1'] = "Date"
    ws['B1'] = "Contract1"
    ws['C1'] = "Close1"
    ws['D1'] = "Contract2"
    ws['E1'] = "Close2"
    ws['F1'] = "Contract3"
    ws['G1'] = "Close3"
    ws['H1'] = "Contract4"
    ws['I1'] = "Close4"
    ws['J1'] = "Contract5"
    ws['K1'] = "Close5"
    ws['L1'] = "Contract6"
    ws['M1'] = "Close6"
    ws['N1'] = "Contract7"
    ws['O1'] = "Close7"
    ws['P1'] = "Contract8"
    ws['Q1'] = "Close8"
    ws['R1'] = "Contract9"
    ws['S1'] = "Close9"
    ws['T1'] = "Contract10"
    ws['U1'] = "Close10"

    row_no=2
    
    for date in x:
        data = result[result[0]==date]
        data=data.reset_index()
        data=data.drop(['index'],axis=1)

        for i in range(0,len(data)):
            ws.cell(row=row_no,column=1).value = data.ix[i,:][0]
            ws.cell(row=row_no,column=int(data.ix[i,:][3])*2).value=data.ix[i,:][1] # save Contact expiry month
            ws.cell(row=row_no,column=(int(data.ix[i,:][3])*2)+1).value=data.ix[i,:][2] # save closing price
    
    row_no = row_no+1

    wb.save("output.xlsx")


#get_daily_vix("localhost","root","test123",os.getcwd())

#store_exceldata_to_db("C:\\Users\\risha\\OneDrive\\Desktop\\Momenta\\Historical_VIX\\yearly data","localhost","root","test123",os.getcwd())
 
#result= get_data_from_db("localhost","root","test123",'vix','vix_data')
    

    